#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Attendance System - Viewer Application
Port: 5000
Password: alusman@123
Read-only access for viewing attendance records
"""

from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file
from functools import wraps
from datetime import datetime, timedelta
import pandas as pd
from io import BytesIO, StringIO
import openpyxl
from openpyxl.styles import Protection, Font, Alignment, PatternFill
from openpyxl.workbook.protection import WorkbookProtection
import os

from config import SECRET_KEY, VIEWER_PASSWORD_HASH, verify_password, VIEWER_PORT, SESSION_TIMEOUT
from utils import (
    get_db, get_date_range, get_attendance_logs, 
    get_all_users, get_total_counts, get_available_floors,
    get_settings
)

app = Flask(__name__)
app.secret_key = SECRET_KEY
app.config['PERMANENT_SESSION_LIFETIME'] = SESSION_TIMEOUT

def get_admin_date_range():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT key, value FROM settings WHERE key IN ('date_from', 'date_to')")
    rows = cursor.fetchall()
    conn.close()
    result = {}
    for row in rows:
        result[row['key']] = row['value']
    return result.get('date_from', ''), result.get('date_to', '')

def get_day_name(date_str):
    try:
        dt = datetime.strptime(date_str, '%Y-%m-%d')
        days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
        return days[dt.weekday()]
    except:
        return ''

def is_sunday(date_str):
    try:
        dt = datetime.strptime(date_str, '%Y-%m-%d')
        return dt.weekday() == 6
    except:
        return False

def is_saturday(date_str):
    try:
        dt = datetime.strptime(date_str, '%Y-%m-%d')
        return dt.weekday() == 5
    except:
        return False

def format_time_str(time_str):
    if not time_str:
        return '--'
    if time_str == '--' or time_str == 'None':
        return '--'
    try:
        if ' ' in time_str:
            time_part = time_str.split(' ')[1]
        else:
            time_part = time_str
        # Check for midnight
        if time_part in ['00:00:00', '12:00:00']:
            return '--'
        hour = int(time_part.split(':')[0])
        minute = time_part.split(':')[1]
        if hour == 0 and minute == '00':
            return '--'
        if hour == 12 and minute == '00':
            return '--'
        ampm = 'AM' if hour < 12 else 'PM'
        hour12 = hour if hour <= 12 else hour - 12
        if hour12 == 0:
            hour12 = 12
        return f"{hour12:02d}:{minute} {ampm}"
    except:
        return '--'

def calculate_remarks(chk_in_time, chk_out_time, date_str):
    if not chk_in_time and not chk_out_time:
        if is_sunday(date_str):
            return '--', '--', 'Weekly Off'
        else:
            return 'Missing In', 'Missing Out', 'Absent/Leave (Document your Truancy)'
    if is_sunday(date_str):
        remark_in = 'Over Time' if chk_in_time else 'Over Time'
        remark_out = 'Over Time' if chk_out_time else 'Over Time'
        return remark_in, remark_out, 'Weekly Off/Over Time'
    is_sat = is_saturday(date_str)
    remark_in = ''
    remark_out = ''
    status = ''
    if chk_in_time and chk_in_time != '--':
        try:
            if ' ' in chk_in_time:
                time_str = chk_in_time.split(' ')[1]
            else:
                time_str = chk_in_time
            hour = int(time_str.split(':')[0])
            minute = int(time_str.split(':')[1])
            total = hour * 60 + minute
            if is_sat:
                if total < 8 * 60:
                    remark_in = 'On Time'
                    status = 'Present'
                elif total <= 10 * 60 + 15:
                    remark_in = 'On Time'
                    status = 'Present'
                elif total <= 11 * 60:
                    remark_in = 'Grace Time'
                    status = 'Present'
                elif total <= 15 * 60 + 30:
                    remark_in = 'Late'
                    status = 'Present'
                else:
                    remark_in = 'Late'
                    status = 'Present'
            else:
                if total < 8 * 60:
                    remark_in = 'On Time'
                    status = 'Present'
                elif total <= 9 * 60 + 15:
                    remark_in = 'On Time'
                    status = 'Present'
                elif total <= 9 * 60 + 30:
                    remark_in = 'Grace Time'
                    status = 'Present'
                elif total <= 11 * 60 + 30:
                    remark_in = 'Late'
                    status = 'Present/Late'
                else:
                    remark_in = 'Half Day'
                    status = 'Present/Half Day'
        except:
            remark_in = 'On Time'
            status = 'Present'
    else:
        remark_in = 'Missing In'
        if chk_out_time:
            status = 'Present/Missing In'
    if chk_out_time and chk_out_time != '--':
        try:
            if ' ' in chk_out_time:
                time_str = chk_out_time.split(' ')[1]
            else:
                time_str = chk_out_time
            hour = int(time_str.split(':')[0])
            minute = int(time_str.split(':')[1])
            total = hour * 60 + minute
            if is_sat:
                if total < 14 * 60:
                    remark_out = 'Early Go'
                    if not status:
                        status = 'Present/Early Go'
                elif total <= 15 * 60 + 30:
                    remark_out = 'On Time'
                    if not status:
                        status = 'Present'
                else:
                    remark_out = 'Over Time'
                    if not status:
                        status = 'Present/Over Time'
            else:
                if total < 17 * 60:
                    remark_out = 'Early Go'
                    if not status:
                        status = 'Present/Early Go'
                elif total <= 18 * 60:
                    remark_out = 'On Time'
                    if not status:
                        status = 'Present'
                else:
                    remark_out = 'Over Time'
                    if not status:
                        status = 'Present/Over Time'
        except:
            remark_out = 'On Time'
            if not status:
                status = 'Present'
    else:
        remark_out = 'Missing Out'
        if chk_in_time and not status:
            status = 'Present/Missing Out'
    return remark_in, remark_out, status

def get_all_dates(date_from, date_to):
    start = datetime.strptime(date_from, '%Y-%m-%d')
    end = datetime.strptime(date_to, '%Y-%m-%d')
    dates = []
    current = start
    while current <= end:
        dates.append(current.strftime('%Y-%m-%d'))
        current += timedelta(days=1)
    return dates

def get_viewer_date_limits():
    admin_from, admin_to = get_admin_date_range()
    today = datetime.now().date()
    today_str = today.strftime('%Y-%m-%d')
    if admin_from and admin_to:
        min_date = admin_from
        max_date = admin_to
    else:
        min_date = (today - timedelta(days=60)).strftime('%Y-%m-%d')
        max_date = today_str
    return min_date, max_date

def get_default_dates():
    today = datetime.now().date()
    date_to = today.strftime('%Y-%m-%d')
    date_from = (today - timedelta(days=6)).strftime('%Y-%m-%d')
    return date_from, date_to

def get_export_data(date_from, date_to, search_user='', selected_floor=''):
    conn = get_db()
    cursor = conn.cursor()
    if search_user:
        cursor.execute("""
            SELECT user_id, name, device_floor FROM users 
            WHERE is_active = 1 AND (user_id LIKE ? OR name LIKE ?)
            ORDER BY device_floor, name
        """, (f'%{search_user}%', f'%{search_user}%'))
    else:
        cursor.execute("""
            SELECT user_id, name, device_floor FROM users 
            WHERE is_active = 1 ORDER BY device_floor, name
        """)
    users = cursor.fetchall()
    if selected_floor:
        users = [u for u in users if u['device_floor'] == selected_floor]
    all_dates = get_all_dates(date_from, date_to)
    data_by_user = {}
    for user in users:
        user_data = []
        for date_str in all_dates:
            day = get_day_name(date_str)
            cursor.execute("""
                SELECT 
                    DATE(timestamp) as date,
                    MIN(CASE WHEN punch_type = 0 THEN timestamp END) as chk_in,
                    MAX(CASE WHEN punch_type = 1 THEN timestamp END) as chk_out,
                    (SELECT remark_in FROM attendance_logs WHERE user_id = ? AND DATE(timestamp) = ? AND device_floor = ? AND punch_type = 0 ORDER BY timestamp ASC LIMIT 1) as remark_in,
                    (SELECT remark_out FROM attendance_logs WHERE user_id = ? AND DATE(timestamp) = ? AND device_floor = ? AND punch_type = 1 ORDER BY timestamp ASC LIMIT 1) as remark_out,
                    (SELECT status_override FROM attendance_logs WHERE user_id = ? AND DATE(timestamp) = ? AND device_floor = ? AND punch_type = 0 ORDER BY timestamp ASC LIMIT 1) as status_override
                FROM attendance_logs 
                WHERE user_id = ? AND DATE(timestamp) = ? AND device_floor = ?
                GROUP BY DATE(timestamp)
            """, (user['user_id'], date_str, user['device_floor'],
                  user['user_id'], date_str, user['device_floor'],
                  user['user_id'], date_str, user['device_floor'],
                  user['user_id'], date_str, user['device_floor']))
            row = cursor.fetchone()
            if row:
                status_override = row['status_override'] or ''
                if status_override:
                    chk_in = '--'
                    chk_out = '--'
                    remark_in = 'Missing In'
                    remark_out = 'Missing Out'
                    status = status_override
                else:
                    chk_in = row['chk_in'] or '--'
                    chk_out = row['chk_out'] or '--'
                    remark_in = row['remark_in'] or 'Missing In'
                    remark_out = row['remark_out'] or 'Missing Out'
                    chk_in = format_time_str(chk_in)
                    chk_out = format_time_str(chk_out)
                    _, _, status = calculate_remarks(row['chk_in'] or '', row['chk_out'] or '', date_str)
            else:
                if day == 'Sunday':
                    chk_in = '--'
                    chk_out = '--'
                    remark_in = '--'
                    remark_out = '--'
                    status = 'Weekly Off'
                else:
                    chk_in = '--'
                    chk_out = '--'
                    remark_in = 'Missing In'
                    remark_out = 'Missing Out'
                    status = 'Absent/Leave (Document your Truancy)'
            user_data.append({
                'Date': date_str,
                'Day': day,
                'Check In': chk_in,
                'Remarks In': remark_in,
                'Check Out': chk_out,
                'Remarks Out': remark_out,
                'Status': status
            })
        data_by_user[(user['user_id'], user['name'], user['device_floor'])] = user_data
    conn.close()
    return data_by_user

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'viewer_logged_in' not in session:
            flash('Please login first', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

@app.route('/')
@app.route('/index')
@login_required
def index():
    default_from, default_to = get_default_dates()
    date_from = request.args.get('date_from', default_from)
    date_to = request.args.get('date_to', default_to)
    min_date, max_date = get_viewer_date_limits()
    if date_from < min_date:
        date_from = min_date
    if date_from > max_date:
        date_from = max_date
    if date_to > max_date:
        date_to = max_date
    if date_to < min_date:
        date_to = min_date
    search_user = request.args.get('user_id', '')
    selected_floor = request.args.get('floor', '')
    conn = get_db()
    cursor = conn.cursor()
    if search_user:
        cursor.execute("""
            SELECT user_id, name, device_floor FROM users 
            WHERE is_active = 1 AND (user_id LIKE ? OR name LIKE ?)
            ORDER BY device_floor, name
        """, (f'%{search_user}%', f'%{search_user}%'))
    else:
        cursor.execute("""
            SELECT user_id, name, device_floor FROM users 
            WHERE is_active = 1 ORDER BY device_floor, name
        """)
    users = cursor.fetchall()
    if selected_floor:
        users = [u for u in users if u['device_floor'] == selected_floor]
    all_dates = get_all_dates(date_from, date_to)
    attendance_lookup = {}
    if users:
        for user in users:
            for date_str in all_dates:
                cursor.execute("""
                    SELECT 
                        al.user_id,
                        al.device_floor,
                        DATE(al.timestamp) as date,
                        MIN(CASE WHEN al.punch_type = 0 THEN al.timestamp END) as chk_in,
                        MAX(CASE WHEN al.punch_type = 1 THEN al.timestamp END) as chk_out,
                        (SELECT al2.remark_in FROM attendance_logs al2 WHERE al2.user_id = al.user_id AND DATE(al2.timestamp) = DATE(al.timestamp) AND al2.device_floor = al.device_floor AND al2.punch_type = 0 ORDER BY al2.timestamp ASC LIMIT 1) as remark_in,
                        (SELECT al2.remark_out FROM attendance_logs al2 WHERE al2.user_id = al.user_id AND DATE(al2.timestamp) = DATE(al.timestamp) AND al2.device_floor = al.device_floor AND al2.punch_type = 1 ORDER BY al2.timestamp ASC LIMIT 1) as remark_out,
                        (SELECT al2.status_override FROM attendance_logs al2 WHERE al2.user_id = al.user_id AND DATE(al2.timestamp) = DATE(al.timestamp) AND al2.device_floor = al.device_floor AND al2.punch_type = 0 ORDER BY al2.timestamp ASC LIMIT 1) as status_override
                    FROM attendance_logs al
                    WHERE al.user_id = ? AND DATE(al.timestamp) = ? AND al.device_floor = ?
                    GROUP BY al.user_id, DATE(al.timestamp), al.device_floor
                """, (user['user_id'], date_str, user['device_floor']))
                row = cursor.fetchone()
                if row:
                    key = (row['user_id'], row['date'])
                    attendance_lookup[key] = row
    conn.close()
    logs = []
    for user in users:
        for date_str in all_dates:
            day = get_day_name(date_str)
            key = (user['user_id'], date_str)
            if key in attendance_lookup:
                row = attendance_lookup[key]
                status_override = row['status_override'] or ''
                chk_in_val = row['chk_in'] or ''
                chk_out_val = row['chk_out'] or ''
                remark_in_final = row['remark_in'] or ''
                remark_out_final = row['remark_out'] or ''
                log_dict = {
                    'user_id': user['user_id'],
                    'device_floor': user['device_floor'],
                    'user_name': user['name'],
                    'date': date_str,
                    'day': day,
                    'chk_in': chk_in_val,
                    'chk_out': chk_out_val,
                    'remark_in_final': remark_in_final,
                    'remark_out_final': remark_out_final,
                    'status_override': status_override,
                    'display_status': ''
                }
                if status_override:
                    log_dict['display_status'] = status_override
                else:
                    if not log_dict['display_status']:
                        remark_in, remark_out, status = calculate_remarks(
                            log_dict['chk_in'], log_dict['chk_out'], date_str
                        )
                        log_dict['remark_in_final'] = log_dict['remark_in_final'] or remark_in
                        log_dict['remark_out_final'] = log_dict['remark_out_final'] or remark_out
                        log_dict['display_status'] = status
            else:
                if day == 'Sunday':
                    log_dict = {
                        'user_id': user['user_id'],
                        'device_floor': user['device_floor'],
                        'user_name': user['name'],
                        'date': date_str,
                        'day': day,
                        'chk_in': '',
                        'chk_out': '',
                        'remark_in_final': '--',
                        'remark_out_final': '--',
                        'status_override': '',
                        'display_status': 'Weekly Off'
                    }
                else:
                    log_dict = {
                        'user_id': user['user_id'],
                        'device_floor': user['device_floor'],
                        'user_name': user['name'],
                        'date': date_str,
                        'day': day,
                        'chk_in': '',
                        'chk_out': '',
                        'remark_in_final': 'Missing In',
                        'remark_out_final': 'Missing Out',
                        'status_override': '',
                        'display_status': 'Absent/Leave (Document your Truancy)'
                    }
            # Format times
            if log_dict['chk_in']:
                log_dict['chk_in'] = format_time_str(str(log_dict['chk_in']))
            else:
                log_dict['chk_in'] = '--'
            if log_dict['chk_out']:
                log_dict['chk_out'] = format_time_str(str(log_dict['chk_out']))
            else:
                log_dict['chk_out'] = '--'
            logs.append(log_dict)
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) as count FROM users WHERE is_active = 1")
    active_users = cursor.fetchone()['count']
    cursor.execute("SELECT DISTINCT device_floor FROM users")
    floors = [row['device_floor'] for row in cursor.fetchall()]
    conn.close()
    settings = get_settings()
    admin_date_from = settings.get('date_from', '')
    admin_date_to = settings.get('date_to', '')
    return render_template('index.html',
                         logs=logs,
                         date_from=date_from,
                         date_to=date_to,
                         search_user=search_user,
                         selected_floor=selected_floor,
                         active_users=active_users,
                         floors=floors,
                         admin_date_from=admin_date_from,
                         admin_date_to=admin_date_to,
                         is_admin=False,
                         viewer_mode=True,
                         min_date=min_date,
                         max_date=max_date)

@app.route('/users')
@login_required
def users():
    users = get_all_users(include_inactive=False)
    return render_template('users.html', users=users, is_admin=False, viewer_mode=True)

@app.route('/reports')
@login_required
def reports():
    return render_template('reports.html', is_admin=False, viewer_mode=True)

@app.route('/activity')
@login_required
def activity():
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    selected_user = request.args.get('user_id', '')
    if not date_from or not date_to:
        today = datetime.now().date()
        date_to = today.strftime('%Y-%m-%d')
        date_from = (today - timedelta(days=6)).strftime('%Y-%m-%d')
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT user_id, name, device_floor FROM users WHERE is_active = 1 ORDER BY device_floor, name")
    all_users = cursor.fetchall()
    all_dates = get_all_dates(date_from, date_to)
    punches = []
    if request.args.get('date_from') or request.args.get('date_to') or request.args.get('user_id'):
        query = """
            SELECT al.user_id, u.name as user_name, al.device_floor,
                   DATE(al.timestamp) as date,
                   MIN(CASE WHEN punch_type = 0 THEN al.timestamp END) as chk_in,
                   MAX(CASE WHEN punch_type = 1 THEN al.timestamp END) as chk_out,
                   CASE strftime('%w', al.timestamp)
                       WHEN '0' THEN 'Sunday' WHEN '1' THEN 'Monday' WHEN '2' THEN 'Tuesday'
                       WHEN '3' THEN 'Wednesday' WHEN '4' THEN 'Thursday' WHEN '5' THEN 'Friday'
                       WHEN '6' THEN 'Saturday'
                   END as day
            FROM attendance_logs al
            LEFT JOIN users u ON al.user_id = u.user_id AND al.device_floor = u.device_floor
            WHERE DATE(al.timestamp) BETWEEN ? AND ?
                AND (al.status_override IS NULL OR al.status_override = '')
        """
        params = [date_from, date_to]
        if selected_user:
            query += " AND al.user_id = ?"
            params.append(selected_user)
        query += " GROUP BY al.user_id, DATE(al.timestamp), al.device_floor"
        query += " ORDER BY DATE(al.timestamp) DESC, al.user_id"
        cursor.execute(query, params)
        combined_punches = cursor.fetchall()
        punch_lookup = {}
        for p in combined_punches:
            key = (p['user_id'], p['date'])
            punch_lookup[key] = p
        if selected_user:
            users_to_show = [u for u in all_users if u['user_id'] == selected_user]
        else:
            users_to_show = all_users
        for user in users_to_show:
            for date_str in all_dates:
                key = (user['user_id'], date_str)
                day = get_day_name(date_str)
                if key in punch_lookup:
                    p = punch_lookup[key]
                    chk_in_time = p['chk_in']
                    chk_out_time = p['chk_out']
                    chk_in_display = format_time_str(chk_in_time)
                    chk_out_display = format_time_str(chk_out_time)
                    punches.append({
                        'user_id': user['user_id'],
                        'device_floor': user['device_floor'],
                        'user_name': user['name'],
                        'date': date_str,
                        'day': day,
                        'time': f"In: {chk_in_display} | Out: {chk_out_display}"
                    })
                else:
                    punches.append({
                        'user_id': user['user_id'],
                        'device_floor': user['device_floor'],
                        'user_name': user['name'],
                        'date': date_str,
                        'day': day,
                        'time': '--'
                    })
    else:
        punches = []
    conn.close()
    return render_template('activity.html',
                         punches=punches, all_users=all_users,
                         date_from=date_from, date_to=date_to,
                         selected_user=selected_user,
                         tab='raw', is_admin=False, viewer_mode=True)

# ============ VIEWER EXPORT ROUTES ============

@app.route('/export/viewer_excel')
@login_required
def export_viewer_excel():
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    search_user = request.args.get('user_id', '')
    selected_floor = request.args.get('floor', '')
    if not date_from or not date_to:
        default_from, default_to = get_default_dates()
        date_from = default_from
        date_to = default_to
    data_by_user = get_export_data(date_from, date_to, search_user, selected_floor)
    if not data_by_user:
        flash('No data to export', 'warning')
        return redirect(url_for('index'))
    output = BytesIO()
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)
    for (user_id, user_name, floor), records in data_by_user.items():
        if records:
            sheet_name = f"{user_name} ({floor.replace('_', ' ')})"[:31]
            ws = wb.create_sheet(title=sheet_name)
            headers = ['Date', 'Day', 'Check In', 'Remarks In', 'Check Out', 'Remarks Out', 'Status']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='1a237e', end_color='1a237e', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
                cell.protection = Protection(locked=True)
            for row_idx, record in enumerate(records, 2):
                ws.cell(row=row_idx, column=1, value=record['Date'])
                ws.cell(row=row_idx, column=2, value=record['Day'])
                ws.cell(row=row_idx, column=3, value=record['Check In'])
                ws.cell(row=row_idx, column=4, value=record['Remarks In'])
                ws.cell(row=row_idx, column=5, value=record['Check Out'])
                ws.cell(row=row_idx, column=6, value=record['Remarks Out'])
                ws.cell(row=row_idx, column=7, value=record['Status'])
                for col in range(1, 8):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.protection = Protection(locked=True)
                    cell.alignment = Alignment(horizontal='center')
            for col in range(1, 8):
                max_length = 15
                for row in range(1, len(records) + 2):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = max_length + 2
            ws.auto_filter.ref = ws.dimensions
            ws.protection.sheet = True
            ws.protection.selectLockedCells = True
            ws.protection.selectUnlockedCells = True
            ws.protection.formatCells = True
            ws.protection.formatColumns = True
            ws.protection.formatRows = True
            ws.protection.insertColumns = False
            ws.protection.insertRows = False
            ws.protection.insertHyperlinks = False
            ws.protection.deleteColumns = False
            ws.protection.deleteRows = False
            ws.protection.sort = True
            ws.protection.autoFilter = True
            ws.protection.pivotTables = False
            ws.protection.objects = False
            ws.protection.scenarios = False
    wb.security = openpyxl.workbook.protection.WorkbookProtection()
    wb.security.lockStructure = True
    wb.security.revisionsPassword = ''
    wb.security.workbookPassword = ''
    wb.security.lockWindows = True
    wb.save(output)
    output.seek(0)
    filename = f"attendance_viewer_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(output, download_name=filename, as_attachment=True,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        password = request.form.get('password', '')
        if verify_password(password, VIEWER_PASSWORD_HASH):
            session.permanent = True
            session['viewer_logged_in'] = True
            session['role'] = 'viewer'
            session['login_time'] = datetime.now().isoformat()
            return redirect(url_for('index'))
        else:
            flash('Invalid password. Please try again.', 'danger')
    return render_template('login.html', role='viewer')

@app.route('/logout')
def logout():
    session.clear()
    flash('Logged out successfully', 'info')
    return redirect(url_for('login'))

@app.errorhandler(404)
def page_not_found(e):
    return render_template('error.html', error='Page not found'), 404

@app.errorhandler(500)
def internal_server_error(e):
    return render_template('error.html', error='Internal server error'), 500

if __name__ == '__main__':
    print("=" * 60)
    print("ATTENDANCE SYSTEM - VIEWER MODE")
    print("=" * 60)
    print(f"URL: http://localhost:{VIEWER_PORT}")
    print(f"Password: alusman@123")
    print("=" * 60)
    print("Starting viewer application...")
    print("Press Ctrl+C to stop")
    print("=" * 60)
    app.run(host='0.0.0.0', port=VIEWER_PORT, debug=True, use_reloader=False)
